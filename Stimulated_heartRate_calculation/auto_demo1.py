import os
from typing import Optional
import json
import h5py
import numpy as np
import neurokit2 as nk
import pandas as pd
from openpyxl import load_workbook, Workbook
from pathlib import Path
import re

def resolve_eeg_mat_path(name: str, source_folder: str, verify_exists: bool = True) -> Optional[str]:
    """
    根据输入的名称生成对应的 .mat 文件路径。
    参数:
        name : str            输入名称 (如 'SLP014_Day2_up')
        source_folder : str   目录路径
        verify_exists : bool  是否检查文件是否真实存在

    返回:
        str | None  对应的文件绝对路径或 None
    """
    if not isinstance(name, str) or not name:
        return None

    parts = name.split('_')
    if len(parts) < 2:
        return None

    subject = parts[0]  # 例如 SLP014

    # Day1 情况
    if parts[1] == 'Day1':
        filename = f"{subject}_base_eeg_data.mat"

    # up / down 情况
    elif len(parts) >= 3 and parts[2] in ('up', 'down'):
        # 如果要去掉首字母 S，改为 subject[1:]
        filename = f"{subject}_{parts[2]}_eeg_data.mat"
    else:
        return None

    full_path = os.path.join(source_folder, filename)

    if verify_exists and not os.path.isfile(full_path):
        return None

    return full_path

def load_dict_from_json(file_path: str) -> dict:
    """
    从 JSON 文件加载字典。

    :param file_path: JSON 文件的路径。
    :return: 加载后的字典，如果文件不存在或出错则返回空字典。
    """
    try:
        with open(file_path, 'r', encoding='utf-8') as json_file:
            data = json.load(json_file)
        print(f"字典已从 {file_path} 加载。")
        return data
    except FileNotFoundError:
        print(f"错误: 文件 '{file_path}' 未找到。")
        return {}
    except json.JSONDecodeError:
        print(f"错误: 文件 '{file_path}' 不是有效的 JSON 格式。")
        return {}

def get_ecg(mat_path):
    """
    输入: mat_path (.mat 文件路径，HDF5格式，包含键 'Totol_data')
    输出: 第7列(索引6)的 ecg_data ndarray
    """
    with h5py.File(mat_path, 'r') as f:
        dataset = f['Totol_data']
        data_np = dataset[:]
        ecg_data = data_np[:, 6]
        return ecg_data

def rpeak_count_table(ecg_segment, sampling_rate, window_seconds=30, step_seconds=30):
    '''
    :function:对心电数据进行分析，30s滑窗，计算R峰的个数、
    :param ecg_segment: 心电数据
    :param sampling_rate: 采样率
    :param window_seconds: 窗宽（以s为单位）
    :param step_seconds: 步长（以秒为单位）
    :return: list
    '''

    window_len = int(window_seconds * sampling_rate)  # 窗口大小
    step_len = int(step_seconds * sampling_rate)  # 步长大小
    n_windows = (len(ecg_segment) - window_len) // step_len + 1  # 窗口数量
    print(n_windows)

    result_r_number = []

    # 滑窗分析
    for w in range(n_windows):
        start = w * step_len
        end = start + window_len
        window_data = ecg_segment[start:end]
        _, rpeak_info = nk.ecg_peaks(window_data, sampling_rate=sampling_rate)
        rpeaks = rpeak_info["ECG_R_Peaks"] # 获取R峰的样本索引（即R峰在数据中的位置）
        R_count = len(rpeaks)
        print(f"第{w}的窗口计算的R峰个数为{R_count}")
        result_r_number.append(R_count)
    return result_r_number

def append_column_to_excel_no_align(filepath, col_name, values, sheet_name="Sheet1"):
    """
    追加一列到现有Excel的指定工作表右侧，不做行对齐。
    - 第1行写表头（col_name）
    - 第2行开始写数据（values）
    - 如果新列更长会自动扩展行，如果更短则该列下方留空
    """
    filepath = Path(filepath)
    values = list(values)

    if filepath.exists():
        wb = load_workbook(filepath) #读取已存在的 Excel 文件（.xlsx 格式）
    else:
        wb = Workbook() #创建新的空白 Excel 文件

    # 获取/创建工作表
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)

    # 计算下一空列（openpyxl的max_column是1基）
    # 若完全空表（A1也为空），从第1列写
    if ws.max_column == 1 and ws.max_row == 1 and ws["A1"].value is None:
        next_col = 1
    else:
        next_col = ws.max_column + 1

    # 写表头（第1行）
    ws.cell(row=1, column=next_col, value=col_name)

    # 从第2行开始写数据
    for i, v in enumerate(values, start=2):
        ws.cell(row=i, column=next_col, value=v)

    wb.save(filepath)
    wb.close()

def transform_name(name: str):
    """
    根据输入名称生成加工后的名称。
    """
    if not isinstance(name, str):
        return None

    parts = name.strip().split('_')
    if len(parts) < 2:
        return None

    # 提取前缀中的数字 (如 SLP014 -> 014 -> 14)
    m = re.match(r'^SLP(\d+)$', parts[0])
    if not m:
        return None
    num_str = m.group(1)
    num_clean = str(int(num_str))  # 去除前导零

    # 情况一：Day1
    if parts[1] == 'Day1':
        return f"{num_clean}_base"

    # 情况二：有第三部分为 up 或 down
    if len(parts) >= 3 and parts[2] in ('up', 'down'):
        return f"{num_clean}_{parts[2]}"

    return None


def header_exists_pd(filepath: str, col_name: str, sheet_name: str = "Sheet1") -> bool:
    """
    使用 pandas 检查 Excel 指定工作表第一行表头是否包含 col_name。
    """
    fp = Path(filepath)
    if not fp.is_file():
        return False

    try:
        # nrows=0 只读取表头 (列名)，不会加载数据内容
        df = pd.read_excel(fp, sheet_name=sheet_name, nrows=0, engine="openpyxl")
    except Exception:
        return False

    # df.columns 就是表头列表
    return str(col_name) in [str(c) for c in df.columns]


# ------------------------------------加载数据-------------------------------------------------------------------
# 加载字典
pairs_dictionary = 'D:\\JetBrains\PyCharm Community Edition 2019.2.4\\project\\SleepPic\\HRV_SleepStage(Zhangsong)\\Stimulated_heartRate_calculation\\pairs_path_1.json'
loaded_dictionary = load_dict_from_json(pairs_dictionary)
# mat源文件路径'E:\\Auditory Sleep Stimulation Data\\MAT_original_filter\\EEG_data'
mat_sourceFolder =
for key_subID, value_pairs in loaded_dictionary.items():

    #print(resolve_eeg_mat_path(key_subID, mat_sourceFolder))
    # --------------------------------------获取心电数据数据-----------------------------------
    mat_path = resolve_eeg_mat_path(key_subID, mat_sourceFolder)
    print(f"正在处理文件为：{mat_path}")
    if mat_path is None:
        continue

    save_file_path = "D:\\研究生\\HR_trend_pro1\\result_R\\R_number.xlsx"
    # 列表头
    # ----------------------------------判断文件是否需要处理----------------------------------------
    col_tilte = transform_name(key_subID)
    isProcess = header_exists_pd(save_file_path, col_tilte, 'firstWave')
    if isProcess == True:
        print("文件夹已经处理           过。")
        continue

    ecg_data = get_ecg(mat_path)
    # 加载标签（这个有异常情况）
    if not value_pairs:# SLP019跳过，后续单独处理
        print(f"{key_subID} 没有任何配对，跳过")
        continue

    start1, end1 = value_pairs[0]
    first_slowWave = ecg_data[int(start1):int(end1)]

    # ------------------------------------处理心电，计算R期数量---------------------------------------
    first_result_r_number = rpeak_count_table(first_slowWave, 1000)
    # ------------------------------------保存数据----------------------------------------------------

    append_column_to_excel_no_align(save_file_path, col_tilte, first_result_r_number, 'firstWave')
    print('第一个慢波保存成功')

    if len(value_pairs) == 1:
        continue

    start2, end2 = value_pairs[1]
    second_slowWave = ecg_data[int(start2):int(end2)]
    second_result_r_number = rpeak_count_table(second_slowWave, 1000)
    append_column_to_excel_no_align(save_file_path, col_tilte, second_result_r_number, 'secondWave')
    print('第二个慢波保存成功')



