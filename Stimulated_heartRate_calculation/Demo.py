# function: 求解刺激时间段内心率的值（以30s为1时间帧，不使用算法，以峰值点数代表心率。）
# author: Zhangsong
# time: 2025-11-14-2025

import os
import mne
import h5py
import numpy as np
import neurokit2 as nk
import pandas as pd
from openpyxl import load_workbook, Workbook
from pathlib import Path

#-----------------------------加载数据------------------------------------------------------------
'''
edf_path = "E:\\Auditory Sleep Stimulation Data\\SLP012_Day3_down\\00点00分\\EEG_filter.edf"
raw = mne.io.read_raw_edf(edf_path, preload=True)
# 获取所有通道数据 (data 是 numpy 数组, times 是时间轴)
data, times = raw[:]
#实验采样率
sampling_rate = raw.info['sfreq']
ecg_data = data[raw.ch_names.index('ECG'), :]
# 获取心电的索引值
ecg_channel_index = raw.ch_names.index('ECG')
print(f"心电的通道编号为{ecg_channel_index}")
'''
# 加载marker数据
marker_path = "E:\\Auditory Sleep Stimulation Data\\3588_data\SLP023_Day1\\23点36分\\points_mark.txt"

'''
结束时间对不上
'''

pairs = []  # 用于保存 (12_time, 15_time)
waiting_for = 12  # 初始状态：先寻找 12
current_12_time = None
with open(marker_path, 'r', encoding='utf-8') as f:
    for line in f:
        line = line.strip()
        if not line:
            continue

        time_str, label_str = line.split(',')
        time = float(time_str)
        label = int(label_str)

        # 状态机逻辑
        if waiting_for == 12:
            if label == 12:
                current_12_time = time
                waiting_for = 15  # 找到12后，下一步找15

        elif waiting_for == 15:
            if label == 15:
                pairs.append((current_12_time, time))
                waiting_for = 12  # 成对结束，再找下一个12

print(pairs)


# 读取mat文件
mat_path = "E:\\Auditory Sleep Stimulation Data\\MAT_original_filter\\EEG_data\\SLP023_base_eeg_data.mat"
with h5py.File(mat_path, 'r') as f:
    # f 的行为类似于一个 Python 字典

    # 1. 查看文件中有哪些变量 (keys)
    print("文件中的变量名:", list(f.keys()))
    if 'Totol_data' in f:
        # 从文件中获取该变量对应的数据集(Dataset)对象
        dataset = f['Totol_data']
        data_np = dataset[:]
        ecg_data = data_np[:, 6]
        print("读取成功")
# 对数据进行切分
start1, end1 = pairs[0]
print(start1, end1)
first_slowWave = ecg_data[int(start1):int(end1)]
start2, end2 = pairs[1]
second_slowWave = ecg_data[int(start2):int(end2)]

# ----------------------------------------计算心率-------------------------------------------------------------
# 函数：计算心率，检测R峰值点，进行存储
def rpeak_count_table(ecg_segment, sampling_rate, window_seconds=30, step_seconds=30):
    '''
    :function:对心电数据进行分析，30s滑窗，计算R峰的个数、
    :param ecg_segment: 心电数据
    :param sampling_rate: 采样率
    :param window_seconds: 窗宽（以s为单位）
    :param step_seconds: 步长（以秒为单位）
    :return: list
    '''

    window_len = int(window_seconds * sampling_rate)  # 窗口大小
    step_len = int(step_seconds * sampling_rate)  # 步长大小
    n_windows = (len(ecg_segment) - window_len) // step_len + 1  # 窗口数量
    print(n_windows)

    result_r_number = []

    # 滑窗分析
    for w in range(n_windows):
        start = w * step_len
        end = start + window_len
        window_data = ecg_data[start:end]
        _, rpeak_info = nk.ecg_peaks(window_data, sampling_rate=sampling_rate)
        rpeaks = rpeak_info["ECG_R_Peaks"] # 获取R峰的样本索引（即R峰在数据中的位置）
        R_count = len(rpeaks)
        print(f"第{w}的窗口计算的R峰个数为{R_count}")
        result_r_number.append(R_count)
    return result_r_number

result_r_number = rpeak_count_table(first_slowWave,1000)
second_r_number = rpeak_count_table(second_slowWave, 1000)

# ---------------------------对数据进行保存----------------------------------------------------------------------
save_file_path = "D:\\研究生\\HR_trend_pro1\\result_R\\R_number.xlsx"
'''
df = pd.DataFrame({"12_down": result_r_number})
with pd.ExcelWriter(save_file_path) as writer:
    df.to_excel(writer,index=False)
'''


def append_column_to_excel_no_align(filepath, col_name, values, sheet_name="firstWave"):
    """
    追加一列到现有Excel的指定工作表右侧，不做行对齐。
    - 第1行写表头（col_name）
    - 第2行开始写数据（values）
    - 如果新列更长会自动扩展行，如果更短则该列下方留空
    """
    filepath = Path(filepath)
    values = list(values)

    if filepath.exists():
        wb = load_workbook(filepath) #读取已存在的 Excel 文件（.xlsx 格式）
    else:
        wb = Workbook() #创建新的空白 Excel 文件

    # 获取/创建工作表
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)

    # 计算下一空列（openpyxl的max_column是1基）
    # 若完全空表（A1也为空），从第1列写
    if ws.max_column == 1 and ws.max_row == 1 and ws["A1"].value is None:
        next_col = 1
    else:
        next_col = ws.max_column + 1

    # 写表头（第1行）
    ws.cell(row=1, column=next_col, value=col_name)

    # 从第2行开始写数据
    for i, v in enumerate(values, start=2):
        ws.cell(row=i, column=next_col, value=v)

    wb.save(filepath)
    wb.close()
append_column_to_excel_no_align(save_file_path,"23_base",result_r_number)
append_column_to_excel_no_align(save_file_path,'23_base', result_r_number, sheet_name='secondWave')

# print(mat_data)





