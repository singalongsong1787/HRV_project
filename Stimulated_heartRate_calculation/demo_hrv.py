# function: 计算数据的心率变异性——同上述的心率数据
# author： Zhangsong
# time: 2025-11-21-1957

import os
import mne
import pandas as pd
import matplotlib.pyplot as plt
from ecgdetectors import Detectors
import neurokit2 as nk
import numpy as np
from hrvanalysis  import get_time_domain_features, get_frequency_domain_features, get_poincare_plot_features
from hrvanalysis.preprocessing import (
    remove_outliers,
    remove_ectopic_beats,
    interpolate_nan_values
)
import h5py
from openpyxl import load_workbook, Workbook
from pathlib import Path

# 加载marker数据
marker_path = "E:\\Auditory Sleep Stimulation Data\\3588_data\SLP023_Day1\\23点36分\\points_mark.txt"

'''
结束时间对不上
'''

pairs = []  # 用于保存 (12_time, 15_time)
waiting_for = 12  # 初始状态：先寻找 12
current_12_time = None
with open(marker_path, 'r', encoding='utf-8') as f:
    for line in f:
        line = line.strip()
        if not line:
            continue

        time_str, label_str = line.split(',')
        time = float(time_str)
        label = int(label_str)

        # 状态机逻辑
        if waiting_for == 12:
            if label == 12:
                current_12_time = time
                waiting_for = 15  # 找到12后，下一步找15

        elif waiting_for == 15:
            if label == 15:
                pairs.append((current_12_time, time))
                waiting_for = 12  # 成对结束，再找下一个12

print(pairs)


def compute_hrv_features(rr_intervals_ms):
    """清理 RR 间期并计算 HRV 特征"""
    # --- 清理 ---

    rr_clean = remove_outliers(rr_intervals_ms, low_rri=300, high_rri=2000)
    rr_clean = remove_ectopic_beats(rr_clean, method="kamath")
    rr_clean = interpolate_nan_values(rr_clean, interpolation_method="linear")

    print(f"rr_clean的数据类型为：{type(rr_clean)}")
    print(f"rr_clean的值为：{rr_clean}")

    # --- HRV 特征 ---
    features = {}
    features.update(get_time_domain_features(rr_intervals_ms))
    return features

# 读取mat文件
mat_path = "E:\\Auditory Sleep Stimulation Data\\MAT_original_filter\\EEG_data\\SLP023_base_eeg_data.mat"
with h5py.File(mat_path, 'r') as f:
    # f 的行为类似于一个 Python 字典

    # 1. 查看文件中有哪些变量 (keys)
    print("文件中的变量名:", list(f.keys()))
    if 'Totol_data' in f:
        # 从文件中获取该变量对应的数据集(Dataset)对象
        dataset = f['Totol_data']
        data_np = dataset[:]
        ecg_data = data_np[:, 6]
        print("读取成功")
print(pairs)
# 对数据进行切分
start1, end1 = pairs[0]
print(start1, end1)
first_slowWave = ecg_data[int(start1):int(end1)]
start2, end2 = pairs[1]
second_slowWave = ecg_data[int(start2):int(end2)]

import numpy as np
import neurokit2 as nk

def hrv_sliding_windows(ecg_data, sampling_rate, window_minutes=5, step_seconds=60):
    """
    使用简单滑窗（固定窗口、固定步长）计算每个窗口的 SDNN 和 RMSSD。

    参数：
    ecg_data：需计算的心电数据
    sampling_rate：采样率
    window_miutes:时间窗
    step_seconds：步长

    返回:
        SDNN_list, RMSSD_list
    """

    window_len = int(window_minutes * 60 * sampling_rate)  # 窗口大小
    step_len = int(step_seconds * sampling_rate)           # 步长大小
    n_windows = (len(ecg_data) - window_len) // step_len + 1

    print(f"窗口数量: {n_windows}")

    SDNN_list = []
    RMSSD_list = []

    for w in range(n_windows):
        start = w * step_len
        end = start + window_len
        ecg_segment = ecg_data[start:end]

        # --- R 峰检测 ---
        _, rpeak_info = nk.ecg_peaks(ecg_segment, sampling_rate=sampling_rate)
        rpeaks = rpeak_info["ECG_R_Peaks"]

        if len(rpeaks) < 2:
            SDNN_list.append(np.nan)
            RMSSD_list.append(np.nan)
            continue

        # --- RR 间期 (ms) ---
        rpeaks_times = np.array(rpeaks) / sampling_rate
        rr_intervals_ms = np.diff(rpeaks_times) * 1000

        # --- 使用你的函数清理 RR + 提取特征 ---
        features = compute_hrv_features(rr_intervals_ms)

        # --- 仅取 SDNN 和 RMSSD ---
        SDNN = features.get("sdnn", np.nan)
        RMSSD = features.get("rmssd", np.nan)

        SDNN_list.append(SDNN)
        RMSSD_list.append(RMSSD)

        print(f"窗口 {w}: SDNN={SDNN:.2f}, RMSSD={RMSSD:.2f}")

    return SDNN_list, RMSSD_list


def append_column_to_excel_no_align(filepath, col_name, values, sheet_name="firstWave"):
    """
    追加一列到现有Excel的指定工作表右侧，不做行对齐。
    - 第1行写表头（col_name）
    - 第2行开始写数据（values）
    - 如果新列更长会自动扩展行，如果更短则该列下方留空
    参数： 保存文件路径//列名称//值//工作区名称

    """
    filepath = Path(filepath)
    values = list(values)

    if filepath.exists():
        wb = load_workbook(filepath) #读取已存在的 Excel 文件（.xlsx 格式）
    else:
        wb = Workbook() #创建新的空白 Excel 文件

    # 获取/创建工作表
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)

    # 计算下一空列（openpyxl的max_column是1基）
    # 若完全空表（A1也为空），从第1列写
    if ws.max_column == 1 and ws.max_row == 1 and ws["A1"].value is None:
        next_col = 1
    else:
        next_col = ws.max_column + 1


    # 写表头（第1行）
    ws.cell(row=1, column=next_col, value=col_name)

    # 从第2行开始写数据
    for i, v in enumerate(values, start=2):
        ws.cell(row=i, column=next_col, value=v)

    wb.save(filepath)
    wb.close()

# #############################计算心率变异性##################################################
sampling_rate = 1000
SDNN_1, RMSSD_1 = hrv_sliding_windows(first_slowWave, sampling_rate)
SDNN_2, RMSSD_2 = hrv_sliding_windows(second_slowWave, sampling_rate)


# ############################ 保存数据到excel文件中 ###########################################
excel_SDNN_path = "D:\\研究生\\HR_trend_pro1\\result_R\\SDNN.xlsx"
excel_RMSSD_path = "D:\\研究生\\HR_trend_pro1\\result_R\\RMSSD.xlsx"
#append_column_to_excel_no_align(excel_SDNN_path, "23_base", SDNN_1, "firstWave")
#append_column_to_excel_no_align(excel_RMSSD_path, "23_base", RMSSD_1, "firstWave")
append_column_to_excel_no_align(excel_SDNN_path, "123_base", SDNN_2, "secondWave")
append_column_to_excel_no_align(excel_RMSSD_path, "23_base", RMSSD_2, "secondWave")







